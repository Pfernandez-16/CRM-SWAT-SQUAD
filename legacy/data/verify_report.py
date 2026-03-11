"""
Verify report numbers for March 2026 against generated data.
Simulates Analytics.js logic in Python for validation.
"""
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook("data/output/CRM_Test_Data.xlsx")

def read_sheet(name):
    ws = wb[name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows

leads = read_sheet("fact_leads")
interacciones = read_sheet("fact_interacciones")
calificacion = read_sheet("fact_calificacion")
contactos = read_sheet("dim_contactos")
deals = read_sheet("fact_deals")

# Build indexes
contactos_idx = {str(c["id_contacto"]): c for c in contactos}
calif_idx = {str(c["id_lead"]): c for c in calificacion}
inter_idx = {}
for inter in interacciones:
    lid = str(inter["id_lead"])
    if lid not in inter_idx:
        inter_idx[lid] = []
    inter_idx[lid].append(inter)

# Date ranges
date_in = datetime(2026, 3, 1)
date_out = datetime(2026, 3, 31)
duration = date_out - date_in  # 30 days
prev_in = date_in - duration
prev_out = date_out - duration

print(f"Current period: {date_in.date()} to {date_out.date()}")
print(f"Previous period: {prev_in.date()} to {prev_out.date()}")

# Filter leads by date range
def filter_leads(leads, d_in, d_out):
    filtered = []
    for lead in leads:
        fi = lead["fecha_ingreso"]
        if fi is None:
            continue
        if isinstance(fi, str):
            fi = datetime.strptime(fi, "%Y-%m-%d %H:%M:%S")
        if fi >= d_in and fi <= d_out:
            filtered.append(lead)
    return filtered

current = filter_leads(leads, date_in, date_out)
previous = filter_leads(leads, prev_in, prev_out)

print(f"\nCurrent leads: {len(current)}")
print(f"Previous leads: {len(previous)}")

# Segment helper
def get_segment(lead):
    cal = calif_idx.get(str(lead["id_lead"]))
    if not cal:
        return "unknown"
    tm = str(cal.get("tipo_membresia") or "").strip()
    if tm == "Manufacturers": return "manufacturers"
    if tm == "Individuals": return "individuals"
    if tm: return "other"
    return "unknown"

# ============ EMBUDO GENERAL ============
print("\n" + "="*60)
print("EMBUDO GENERAL")
print("="*60)

def count_metric(leads_list, test_fn):
    total = mfg = ind = 0
    for lead in leads_list:
        if test_fn(lead):
            total += 1
            seg = get_segment(lead)
            if seg == "manufacturers": mfg += 1
            elif seg == "individuals": ind += 1
    return total, mfg, ind

def calc_delta(cur, prev):
    if prev == 0 and cur > 0: return "+100.0%"
    if prev == 0 and cur == 0: return "0%"
    return f"{((cur - prev) / prev * 100):+.1f}%"

def print_metric(name, cur_leads, prev_leads, test_fn):
    ct, cm, ci = count_metric(cur_leads, test_fn)
    pt, pm, pi = count_metric(prev_leads, test_fn)
    print(f"\n{name}:")
    print(f"  Total:  {ct} (prev: {pt}, delta: {calc_delta(ct, pt)})")
    print(f"  Mfg:    {cm} ({cm/ct*100:.1f}% of total, prev: {pm}, delta: {calc_delta(cm, pm)})" if ct > 0 else f"  Mfg:    {cm}")
    print(f"  Ind:    {ci} ({ci/ct*100:.1f}% of total, prev: {pi}, delta: {calc_delta(ci, pi)})" if ct > 0 else f"  Ind:    {ci}")

# 1. Total Leads
print_metric("1. Total Leads", current, previous, lambda l: True)

# 2. Contactables
def is_contactable(lead):
    c = contactos_idx.get(str(lead["id_contacto"]))
    if not c: return False
    return bool(str(c.get("telefono_1") or "").strip()) or bool(str(c.get("email") or "").strip())
print_metric("2. Contactables", current, previous, is_contactable)

# 3. Contactados
def is_contactado(lead):
    return len(inter_idx.get(str(lead["id_lead"]), [])) > 0
print_metric("3. Contactados", current, previous, is_contactado)

# 4. Con Respuesta
def has_contesto(lead):
    inters = inter_idx.get(str(lead["id_lead"]), [])
    return any(str(i.get("resultado") or "").strip() == "Contesto" for i in inters)
print_metric("4. Con Respuesta", current, previous, has_contesto)

# 5. Dialogo Completo
def has_consecutive_contesto(lead):
    inters = inter_idx.get(str(lead["id_lead"]), [])
    toques = sorted([int(i["numero_toque"]) for i in inters if str(i.get("resultado") or "").strip() == "Contesto"])
    if len(toques) < 2: return False
    for j in range(1, len(toques)):
        if toques[j] - toques[j-1] == 1: return True
    return False
print_metric("5. Dialogo Completo", current, previous, has_consecutive_contesto)

# 6. Dialogo Intermitente
def is_intermitente(lead):
    return has_contesto(lead) and not has_consecutive_contesto(lead)
print_metric("6. Dialogo Intermitente", current, previous, is_intermitente)

# 7. Interes
def has_interes(lead):
    cal = calif_idx.get(str(lead["id_lead"]))
    if not cal: return False
    v = str(cal.get("mostro_interes_genuino") or "").strip()
    return v in ("Si", "Sí")
print_metric("7. Interes", current, previous, has_interes)

# 8. Descartados
def is_perdido(lead):
    return str(lead.get("status") or "").strip() == "Perdido"
print_metric("8. Descartados", current, previous, is_perdido)

# 9. Asignados a Ventas
def is_paso_ventas(lead):
    return str(lead.get("status") or "").strip() == "Paso a Ventas"
print_metric("9. Asignados a Ventas", current, previous, is_paso_ventas)

# 10. Deals Cerrados (Vendido)
print("\n--- Deal metrics ---")
cur_lead_ids = {str(l["id_lead"]) for l in current}
prev_lead_ids = {str(l["id_lead"]) for l in previous}

cur_deals = [d for d in deals if str(d["id_lead"]) in cur_lead_ids]
prev_deals = [d for d in deals if str(d["id_lead"]) in prev_lead_ids]

cur_vendidos = [d for d in cur_deals if str(d.get("status_venta") or "").strip() == "Vendido"]
prev_vendidos = [d for d in prev_deals if str(d.get("status_venta") or "").strip() == "Vendido"]

print(f"\nDeals (current leads): {len(cur_deals)}")
print(f"Deals (previous leads): {len(prev_deals)}")
print(f"Vendidos current: {len(cur_vendidos)}, previous: {len(prev_vendidos)}")

cur_cierre = sum(float(d.get("monto_cierre") or 0) for d in cur_vendidos)
prev_cierre = sum(float(d.get("monto_cierre") or 0) for d in prev_vendidos)
print(f"Monto cierres current: ${cur_cierre:,.0f}, previous: ${prev_cierre:,.0f}")

# ============ RAZONES PERDIO LA VENTA ============
print("\n" + "="*60)
print("RAZONES PERDIO LA VENTA")
print("="*60)

cur_perdidos_deals = [d for d in cur_deals if str(d.get("status_venta") or "").strip() == "Perdido"]
prev_perdidos_deals = [d for d in prev_deals if str(d.get("status_venta") or "").strip() == "Perdido"]

print(f"\nDeals Perdidos current: {len(cur_perdidos_deals)}")
print(f"Deals Perdidos previous: {len(prev_perdidos_deals)}")

print("\nRazones in current lost deals:")
for d in cur_perdidos_deals:
    razon = str(d.get("razon_perdida") or "").strip()
    lead_id = d["id_lead"]
    seg = "?"
    for l in current:
        if str(l["id_lead"]) == str(lead_id):
            seg = get_segment(l)
            break
    print(f"  Deal {d['id_deal']}: lead={lead_id}, razon='{razon}', segment={seg}")

print("\nRazones in previous lost deals:")
for d in prev_perdidos_deals:
    razon = str(d.get("razon_perdida") or "").strip()
    lead_id = d["id_lead"]
    seg = "?"
    for l in previous:
        if str(l["id_lead"]) == str(lead_id):
            seg = get_segment(l)
            break
    print(f"  Deal {d['id_deal']}: lead={lead_id}, razon='{razon}', segment={seg}")

# Check accent mismatches in razones
print("\n--- Accent check in razonesMap ---")
code_labels = [
    'Sin presupuesto', 'Eligio competidor', 'No responde', 'Timing inadecuado',
    'No tiene poder de decision', 'Producto no se ajusta', 'Cambio de prioridades',
    'Mala experiencia previa', 'Precio muy alto', 'Proceso interno largo',
    'Se fue con otra solucion', 'No era el perfil', 'Empresa cerro'
]
data_razones = set()
for d in deals:
    r = str(d.get("razon_perdida") or "").strip()
    if r:
        data_razones.add(r)

print(f"Unique razones in data: {data_razones}")
print(f"Labels in code: {set(code_labels)}")
print(f"In data but NOT in code labels: {data_razones - set(code_labels)}")
print(f"In code but NOT in data: {set(code_labels) - data_razones}")

# ============ INCONTACTABLES ============
print("\n" + "="*60)
print("INCONTACTABLES")
print("="*60)

def is_duplicado(lead):
    return str(lead.get("status") or "").strip() == "Duplicado"
def is_invalido(lead):
    s = str(lead.get("status") or "").strip()
    return s in ("Invalido", "Inválido")

print_metric("Duplicado", current, previous, is_duplicado)
print_metric("Equivocado (Inválido)", current, previous, is_invalido)

# Status distribution
print("\n" + "="*60)
print("STATUS DISTRIBUTION")
print("="*60)
print("\nCurrent (Mar):")
status_cur = {}
for l in current:
    s = str(l.get("status") or "").strip()
    status_cur[s] = status_cur.get(s, 0) + 1
for s, c in sorted(status_cur.items(), key=lambda x: -x[1]):
    print(f"  {s}: {c}")

print("\nPrevious:")
status_prev = {}
for l in previous:
    s = str(l.get("status") or "").strip()
    status_prev[s] = status_prev.get(s, 0) + 1
for s, c in sorted(status_prev.items(), key=lambda x: -x[1]):
    print(f"  {s}: {c}")

wb.close()
