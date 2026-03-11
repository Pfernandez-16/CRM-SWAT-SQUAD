"""
Generate CRM_Test_Data.xlsx — identical structure to CRM_Normalizado_v6.xlsx
but with fictional data for Jan-Mar 2026.

Copies reference sheets as-is, replaces fact tables + dim_contactos with generated data.
"""

import random
import os
from datetime import datetime, timedelta
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter

random.seed(42)

BASE_DIR = os.path.dirname(__file__)
SOURCE = os.path.join(BASE_DIR, "CRM_Normalizado_v6.xlsx")
OUTPUT = os.path.join(BASE_DIR, "output", "CRM_Test_Data.xlsx")

# Sheets to copy as-is (reference data)
COPY_SHEETS = [
    "MODELO_DATOS", "CAMBIOS_VS_ORIGINAL", "dim_campanas",
    "dim_vendedores", "dim_productos", "cat_opciones", "config_users"
]

# Sheets to replace with generated data
REPLACE_SHEETS = [
    "dim_contactos", "fact_leads", "fact_deals",
    "fact_interacciones", "fact_calificacion", "log_transacciones"
]

# ============ CONSTANTS ============

NOMBRES = [
    "Santiago", "Valentina", "Mateo", "Isabella", "Sebastián", "Camila", "Alejandro",
    "Sofía", "Daniel", "Mariana", "Andrés", "Luciana", "Diego", "Gabriela", "Nicolás",
    "Carolina", "Fernando", "Paula", "Ricardo", "Andrea", "Miguel", "Laura", "Carlos",
    "María", "Eduardo", "Ana", "Roberto", "Claudia", "Francisco", "Diana", "Jorge",
    "Mónica", "Luis", "Patricia", "Rafael", "Natalia", "Javier", "Catalina", "Tomás",
    "Fernanda", "Emilio", "Daniela", "Hugo", "Jimena", "Pablo", "Lorena", "Martín",
    "Adriana", "Gabriel", "Verónica", "Manuel", "Beatriz", "Oscar", "Rosa", "Alberto",
    "Carmen", "Raúl", "Gloria", "Alfredo", "Teresa", "Esteban", "Silvia", "Iván",
    "Marta", "Germán", "Alicia", "Joaquín", "Elena", "Ramón", "Cecilia"
]

APELLIDOS = [
    "García", "Rodríguez", "Martínez", "López", "González", "Hernández", "Pérez",
    "Sánchez", "Ramírez", "Torres", "Flores", "Rivera", "Gómez", "Díaz", "Cruz",
    "Morales", "Reyes", "Gutiérrez", "Ortiz", "Ramos", "Vargas", "Castillo", "Jiménez",
    "Rojas", "Aguilar", "Mendoza", "Medina", "Castro", "Herrera", "Vega", "Ruiz",
    "Navarro", "Delgado", "Ponce", "Salazar", "Guerrero", "Campos", "Ríos", "Figueroa",
    "Acosta", "Molina", "Silva", "Domínguez", "Cortés", "Paredes", "Estrada", "Valenzuela",
    "León", "Espinoza", "Montes", "Sandoval", "Ochoa", "Miranda", "Fuentes", "Contreras"
]

EMPRESAS = [
    "Grupo Industrial MX", "TechLatam Solutions", "Fabricaciones del Sur", "Innovación Digital",
    "Metal Works SA", "Textiles Modernos", "Consultoría Integral", "Alimentos del Pacífico",
    "Construcciones Unidas", "Logística Express", "Corporativo Norte", "Manufactura Premium",
    "Green Energy LATAM", "BioFarma Continental", "Automotriz Central", "Electrónica Global",
    "Minera del Valle", "Agroindustrias Real", "Plásticos Avanzados", "Química Especializada",
    "TransCargo SA", "Aceros Industriales", "Papel y Cartón MX", "Maquinaria Pesada",
    "Cemento del Golfo", "Vidrio y Cristal", "Aluminios del Norte", "Empaques Sustentables",
    "Textil Orgánico", "Pinturas Internacionales", "Caucho y Derivados", "Envases PET",
    "Productos Lácteos", "Bebidas del Valle", "Snacks Premium", "Fertilizantes del Campo",
    "Software Empresarial", "Cloud Services MX", "Cybersecurity Corp", "Fintech Soluciones",
    "EdTech Innovación", "HealthTech Partners", "PropTech Desarrollo", "AgriTech Solutions",
    "RetailTech Global", "MediaTech Creative", "Industrias Omega", "Corporativo Atlas",
    "Grupo Vanguardia", "Soluciones Integradas"
]

AREAS = [None, "Ventas", "Marketing", "RRHH", "Operaciones", "TI", "Finanzas", "Dirección", "Compras", "Producción"]
PAISES_CIUDADES = [
    ("México", ["CDMX", "Monterrey", "Guadalajara", "Puebla", "Querétaro", "Tijuana", "León"]),
    ("Colombia", ["Bogotá", "Medellín", "Cali", "Barranquilla", "Cartagena"]),
    ("Argentina", ["Buenos Aires", "Córdoba", "Rosario", "Mendoza"]),
    ("Chile", ["Santiago", "Valparaíso", "Concepción"]),
    ("Perú", ["Lima", "Arequipa", "Trujillo"]),
    ("Costa Rica", ["San José"]),
    ("Uruguay", ["Montevideo"]),
    ("Panamá", ["Ciudad de Panamá"]),
    ("Ecuador", ["Quito", "Guayaquil"]),
]
EMPLEADOS_OPTS = [None, "1-10", "11-50", "51-200", "201-500", "501-1000", "1000+"]
NIVELES = ["CEO", "VP", "Director", "Gerente", "Coordinador", "Analista", "Especialista", "Jefe de Área"]

STATUS_WEIGHTS = {
    "jan": {"Nuevo": 5, "Contactado": 10, "En Seguimiento": 18, "Calificado": 12,
            "No Calificado": 10, "Contactado sin Respuesta": 5, "Paso a Ventas": 8,
            "Perdido": 25, "Duplicado": 4, "Inválido": 3},
    "feb": {"Nuevo": 4, "Contactado": 9, "En Seguimiento": 20, "Calificado": 16,
            "No Calificado": 8, "Contactado sin Respuesta": 4, "Paso a Ventas": 12,
            "Perdido": 20, "Duplicado": 4, "Inválido": 3},
    "mar": {"Nuevo": 6, "Contactado": 8, "En Seguimiento": 18, "Calificado": 20,
            "No Calificado": 6, "Contactado sin Respuesta": 3, "Paso a Ventas": 16,
            "Perdido": 16, "Duplicado": 4, "Inválido": 3},
}

CALIDAD_WEIGHTS = {
    "jan": {"Excelente": 10, "Bueno": 20, "Regular": 30, "Malo": 25, "Sin Calificar": 15},
    "feb": {"Excelente": 15, "Bueno": 25, "Regular": 25, "Malo": 20, "Sin Calificar": 15},
    "mar": {"Excelente": 20, "Bueno": 30, "Regular": 25, "Malo": 15, "Sin Calificar": 10},
}

SERVICIO_INTERES = [None, "Membresía Básica", "Membresía Premium", "Membresía Enterprise", "Consultoría", "Capacitación", "Software CRM"]
TIPO_SEGUIMIENTO = ["Llamada", "WhatsApp", "Email", "Reunión presencial", "Videollamada"]
STATUS_SEGUIMIENTO = ["Activo", "Pausado", "Cerrado", "Cancelado", "Esperando respuesta"]
TIPO_INTERACCION = ["Telefono", "WhatsApp", "Correo"]
RESULTADO_INTERACCION = ["Contesto", "No Contesto"]
NOTAS_LEAD = [
    "Agenda reunión próxima semana", "Sin interés en este momento", "Pidió propuesta formal",
    "Muy interesado, buen perfil", "Quiere demo del producto", "Presupuesto limitado",
    "Necesita aprobación de directivos", "Comparando con competidores", "Pide más información",
    "Callback la próxima semana", "Interesado en membership premium", "No contesta llamadas",
    "Respondió por WhatsApp, positivo", "Quiere reunión presencial", "Esperando presupuesto Q2",
    "Excelente prospecto", "Referido por cliente actual", "Necesita cotización personalizada",
    None, None, None, None
]

NECESIDAD_OPTS = ["Necesita CRM", "Busca membresía", "Quiere capacitación", "Busca networking",
                  "Necesita certificación", "Quiere expandir mercado", "Busca proveedores"]
MONTO_PRESUPUESTO_OPTS = ["< $5,000", "$5,000-$20,000", "$20,000-$50,000", "> $50,000"]
SI_NO_PARCIAL = ["Sí", "No", "Parcialmente"]
REGIONES = ["Norteamérica", "Latinoamérica", "Europa", "Asia-Pacífico"]
TIPO_MEMBRESIA = ["Manufacturers", "Individuals", "Attraction"]
MEMBRESIA_WEIGHTS = [55, 30, 15]

DEAL_STATUS_WEIGHTS = {
    "jan": {"Vendido": 8, "Perdido": 20, "En negociación": 15, "Propuesta enviada": 12,
            "Apartado": 10, "Stand-by": 10, "Contactado": 8, "Recien llegado": 5,
            "Reagendado": 5, "No contesta": 4, "No interesado": 3},
    "feb": {"Vendido": 12, "Perdido": 16, "En negociación": 18, "Propuesta enviada": 14,
            "Apartado": 12, "Stand-by": 8, "Contactado": 6, "Recien llegado": 4,
            "Reagendado": 4, "No contesta": 3, "No interesado": 3},
    "mar": {"Vendido": 18, "Perdido": 14, "En negociación": 20, "Propuesta enviada": 15,
            "Apartado": 14, "Stand-by": 6, "Contactado": 4, "Recien llegado": 3,
            "Reagendado": 3, "No contesta": 2, "No interesado": 1},
}
PROYECCION_OPTS = ["Alta", "Media", "Baja"]
RAZON_PERDIDA = [
    "Sin presupuesto", "Eligió competidor", "No responde", "Timing inadecuado",
    "No tiene poder de decisión", "Producto no se ajusta", "Cambio de prioridades",
    "Mala experiencia previa", "Precio muy alto", "Proceso interno largo",
    "Se fue con otra solución", "No era el perfil", "Empresa cerró"
]
RAZON_PERDIDA_WEIGHTS = [15, 12, 10, 8, 7, 8, 6, 3, 10, 5, 8, 5, 3]
PRODUCTO_CIERRE = ["1 Year Membership", "2 Year Membership", "Premium Package", "Enterprise Bundle"]
FUENTE_ORIGEN = ["Inbound", "Outbound", "Referido", "Evento", "Partner", "Webinar"]
FUENTE_WEIGHTS = [30, 25, 15, 12, 10, 8]
TIPO_TRANSACCION = ["Nueva venta", "Renovación", "Upgrade", "Cross-sell", "Re-compra"]
TIPO_TRANSACCION_WEIGHTS = [40, 20, 15, 15, 10]
NOTAS_VENDEDOR = [
    "Muy interesado, buen perfil", "Compara con 2 competidores", "Presupuesto limitado este trimestre",
    "Quiere cerrar antes de fin de mes", "Necesita aprobación de junta", "En proceso de evaluación interna",
    "Pidió descuento adicional", "Interesado en paquete premium", "Referido de cliente satisfecho",
    "Empresa en crecimiento, buen momento", "Tiene urgencia por certificación", "Esperando respuesta de dirección",
    None, None
]

VENDEDOR_EMAILS = [
    "pedro.mendez@swatsquad.com", "ana.garcia@swatsquad.com",
    "carlos.lopez@swatsquad.com", "sofia.martinez@swatsquad.com",
    "maria.torres@swatsquad.com", "roberto.silva@swatsquad.com",
    "lucia.herrera@swatsquad.com"
]


# ============ HELPERS ============

def weighted_choice(options_weights):
    options = list(options_weights.keys())
    weights = list(options_weights.values())
    return random.choices(options, weights=weights, k=1)[0]

def weighted_choice_list(options, weights):
    return random.choices(options, weights=weights, k=1)[0]

def random_date(start, end):
    delta = end - start
    secs = random.randint(0, max(int(delta.total_seconds()), 1))
    return start + timedelta(seconds=secs)

def random_phone():
    prefix = random.choice(["+52", "+57", "+54", "+56", "+51", "+506", "+598", "+507", "+593"])
    return f"{prefix} {random.randint(100,999)} {random.randint(100,999)} {random.randint(1000,9999)}"

def month_key(dt):
    return {1: "jan", 2: "feb", 3: "mar"}.get(dt.month, "mar")

def strip_accents(s):
    return s.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ñ','n')


# ============ DATA GENERATORS ============

def gen_contactos(count):
    rows = []
    used_emails = set()
    for i in range(1, count + 1):
        nombre = random.choice(NOMBRES)
        apellido = random.choice(APELLIDOS)
        base = f"{strip_accents(nombre.lower())}.{strip_accents(apellido.lower())}"
        domain = f"{'corporativo' if random.random() < 0.4 else 'empresa'}.{'mx' if random.random() < 0.5 else 'com'}"
        email = f"{base}@{domain}"
        sfx = 1
        while email in used_emails:
            email = f"{base}{sfx}@empresa.com"
            sfx += 1
        used_emails.add(email)

        pais, ciudades = random.choice(PAISES_CIUDADES)
        rows.append([
            i, nombre, apellido, email, random_phone(),
            random_phone() if random.random() < 0.15 else None,
            random.choice(EMPRESAS), random.choice(AREAS),
            pais, random.choice(ciudades), random.choice(EMPLEADOS_OPTS),
            random.choice(NIVELES),
            random_date(datetime(2025, 12, 1), datetime(2026, 3, 4))
        ])
    return rows


def gen_leads(count_per_month):
    rows = []
    meta = []
    ranges = {
        "jan": (datetime(2026, 1, 1), datetime(2026, 1, 31, 23, 59, 59)),
        "feb": (datetime(2026, 2, 1), datetime(2026, 2, 28, 23, 59, 59)),
        "mar": (datetime(2026, 3, 1), datetime(2026, 3, 4, 23, 59, 59)),
    }
    lid = 0
    for month, count in count_per_month.items():
        start, end = ranges[month]
        for _ in range(count):
            lid += 1
            status = weighted_choice(STATUS_WEIGHTS[month])
            calidad = weighted_choice(CALIDAD_WEIGHTS[month])
            fecha_ing = random_date(start, end)
            fecha_asig = fecha_ing + timedelta(hours=random.randint(6, 72))

            if status in ("Nuevo",): nt = 0
            elif status in ("Contactado", "Duplicado", "Inválido"): nt = random.randint(1, 2)
            elif status in ("En Seguimiento", "Contactado sin Respuesta"): nt = random.randint(2, 5)
            elif status in ("Calificado", "No Calificado"): nt = random.randint(3, 6)
            elif status == "Paso a Ventas": nt = random.randint(3, 7)
            elif status == "Perdido": nt = random.randint(1, 8)
            else: nt = random.randint(1, 4)

            fecha_ult = fecha_ing + timedelta(days=random.randint(1, 30)) if nt > 0 else None

            rows.append([
                lid, lid, random.randint(1, 30), random.randint(1, 5),
                status, calidad, random.choice(SERVICIO_INTERES),
                fecha_ing, fecha_asig, fecha_ult, nt,
                random.choice(TIPO_SEGUIMIENTO) if nt > 0 else None,
                random.choice(STATUS_SEGUIMIENTO) if nt > 0 else None,
                random.choice(NOTAS_LEAD)
            ])
            meta.append({
                "id_lead": lid, "id_contacto": lid, "status": status,
                "calidad": calidad, "fecha_ingreso": fecha_ing,
                "fecha_asignacion": fecha_asig, "fecha_ultimo": fecha_ult,
                "numero_toques": nt, "month": month,
            })
    return rows, meta


def gen_interacciones(meta):
    rows = []
    iid = 0
    ch_w = [35, 45, 20]
    for lead in meta:
        if lead["numero_toques"] == 0:
            continue
        for toque in range(1, lead["numero_toques"] + 1):
            iid += 1
            tipo = weighted_choice_list(TIPO_INTERACCION, ch_w)
            base_prob = 0.40
            if lead["status"] in ("Calificado", "Paso a Ventas"): base_prob = 0.55
            elif lead["status"] in ("Perdido", "No Calificado"): base_prob = 0.30
            elif lead["status"] == "Contactado sin Respuesta": base_prob = 0.10
            prob = max(base_prob - (toque - 1) * 0.04, 0.08)
            resultado = "Contesto" if random.random() < prob else "No Contesto"

            if lead["fecha_ultimo"]:
                ts = random_date(lead["fecha_ingreso"] + timedelta(hours=toque * 12),
                                 lead["fecha_ultimo"] + timedelta(days=1))
            else:
                ts = lead["fecha_ingreso"] + timedelta(hours=toque * random.randint(12, 48))

            dur = random.randint(30, 600) if tipo == "Telefono" and resultado == "Contesto" else None
            nota = None
            if resultado == "Contesto" and random.random() < 0.3:
                nota = random.choice(["Buena conversación, agenda demo", "Interesado, pide más info",
                                      "Confirmó interés", "Quiere cotización", "Disponible siguiente semana"])
            elif resultado == "No Contesto" and random.random() < 0.15:
                nota = random.choice(["Buzón de voz", "No disponible", "Línea ocupada", "Sin respuesta"])

            rows.append([iid, lead["id_lead"], None, random.randint(1, 5), tipo, resultado,
                         toque, ts, dur, nota])
    return rows


def gen_calificacion(meta):
    rows = []
    cid = 0
    eligible = {"Calificado", "Paso a Ventas", "Perdido", "En Seguimiento", "No Calificado"}
    for lead in meta:
        if lead["status"] not in eligible or random.random() < 0.15:
            continue
        cid += 1
        if lead["status"] in ("Calificado", "Paso a Ventas"):
            entendio = random.choice(["Sí", "Sí", "Sí", "Parcialmente"])
            interes = random.choice(["Sí", "Sí", "Sí", "Parcialmente"])
            perfil = random.choice(["Sí", "Sí", "Sí", "No"])
            tercero = random.choice(["No", "No", "Parcialmente", "Sí"])
            presupuesto = random.choice(["Sí", "Sí", "Parcialmente", "No"])
        elif lead["status"] == "Perdido":
            entendio = random.choice(["No", "Parcialmente", "Sí", "No"])
            interes = random.choice(["No", "No", "Parcialmente", "Sí"])
            perfil = random.choice(["No", "No", "Sí", "Parcialmente"])
            tercero = random.choice(["Sí", "Sí", "No", "Parcialmente"])
            presupuesto = random.choice(["No", "No", "Parcialmente", "Sí"])
        else:
            entendio = random.choice(SI_NO_PARCIAL)
            interes = random.choice(SI_NO_PARCIAL)
            perfil = random.choice(["Sí", "No"])
            tercero = random.choice(SI_NO_PARCIAL)
            presupuesto = random.choice(SI_NO_PARCIAL)

        monto = random.choice(MONTO_PRESUPUESTO_OPTS) if presupuesto != "No" else random.choice(["< $5,000", "< $5,000", "$5,000-$20,000"])
        rows.append([
            cid, lead["id_lead"], entendio, interes,
            random.choice(NECESIDAD_OPTS), perfil, tercero, presupuesto, monto,
            random.choice(SI_NO_PARCIAL + [None]),
            random.choice(REGIONES),
            weighted_choice_list(TIPO_MEMBRESIA, MEMBRESIA_WEIGHTS),
            lead["fecha_ingreso"] + timedelta(days=random.randint(3, 14))
        ])
    return rows


def gen_deals(meta):
    rows = []
    did = 0
    for lead in meta:
        make = lead["status"] == "Paso a Ventas" or (lead["status"] == "Calificado" and random.random() < 0.25)
        if not make:
            continue
        did += 1
        sv = weighted_choice(DEAL_STATUS_WEIGHTS[lead["month"]])
        mp = random.randint(2000, 50000)
        ma = int(mp * random.choice([0, 0, 0.1, 0.2, 0.3])) if sv in ("Apartado", "Vendido") else 0
        mc = int(mp * random.uniform(0.7, 1.0)) if sv == "Vendido" else 0
        fp = lead["fecha_asignacion"] + timedelta(days=random.randint(0, 5))
        fpa = fp + timedelta(hours=random.randint(4, 48))
        fc = fp + timedelta(days=random.randint(7, 30)) if sv in ("Vendido", "Perdido") else None
        razon = weighted_choice_list(RAZON_PERDIDA, RAZON_PERDIDA_WEIGHTS) if sv == "Perdido" else None
        rows.append([
            did, lead["id_lead"], lead["id_contacto"],
            random.choice([6, 7]), random.randint(1, 60),
            sv, random.choice(PROYECCION_OPTS), mp, ma, mc,
            fp, fpa, fc, razon,
            random.choice([0, 0, 0, 5, 10, 15, 20]),
            random.random() < 0.08, random.random() < 0.05,
            random.choice(PRODUCTO_CIERRE) if sv == "Vendido" else None,
            weighted_choice_list(FUENTE_ORIGEN, FUENTE_WEIGHTS),
            weighted_choice_list(TIPO_TRANSACCION, TIPO_TRANSACCION_WEIGHTS),
            random.choice(NOTAS_VENDEDOR)
        ])
    return rows


def gen_log(meta, deals):
    rows = []
    lid = 0
    for lead in meta:
        transitions = []
        s = lead["status"]
        if s in ("Contactado", "En Seguimiento", "Calificado", "Paso a Ventas"):
            transitions.append(("Status", "Nuevo", "Contactado"))
            if s in ("En Seguimiento", "Calificado", "Paso a Ventas"):
                transitions.append(("Status", "Contactado", "En Seguimiento"))
            if s in ("Calificado", "Paso a Ventas"):
                transitions.append(("Status", "En Seguimiento", "Calificado"))
            if s == "Paso a Ventas":
                transitions.append(("Status", "Calificado", "Paso a Ventas"))
        elif s == "Perdido":
            transitions.append(("Status", "Nuevo", "Contactado"))
            transitions.append(("Status", "Contactado", "Perdido"))
        elif s in ("Duplicado", "Inválido"):
            transitions.append(("Status", "Nuevo", s))
        elif s == "No Calificado":
            transitions.append(("Status", "Nuevo", "Contactado"))
            transitions.append(("Status", "Contactado", "No Calificado"))
        elif s == "Contactado sin Respuesta":
            transitions.append(("Status", "Nuevo", "Contactado sin Respuesta"))

        for j, (campo, ant, nue) in enumerate(transitions):
            lid += 1
            ts = lead["fecha_ingreso"] + timedelta(hours=random.randint(12, 72) * (j + 1))
            rows.append([lid, ts, "Lead", lead["id_lead"],
                         random.choice(VENDEDOR_EMAILS), campo, ant, nue])

        if random.random() < 0.3 and lead["calidad"] != "Sin Calificar":
            lid += 1
            ts = lead["fecha_ingreso"] + timedelta(days=random.randint(2, 10))
            rows.append([lid, ts, "Lead", lead["id_lead"],
                         random.choice(VENDEDOR_EMAILS), "Calidad de Contacto",
                         "Sin Calificar", lead["calidad"]])

    for deal in deals:
        lid += 1
        rows.append([lid, deal[10], "Deal", deal[0],
                     random.choice(VENDEDOR_EMAILS), "Status Venta",
                     "Recien llegado", deal[5]])
    return rows


# ============ XLSX BUILDER ============

def copy_sheet(src_ws, dst_ws):
    """Copy all data and basic formatting from source to destination worksheet."""
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width

    # Copy merged cells
    for merge in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge))


def write_sheet(ws, headers, data_rows):
    """Write headers + data to worksheet."""
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Auto-width columns
    for c in range(1, len(headers) + 1):
        max_len = len(str(headers[c-1]))
        for r in range(2, min(len(data_rows) + 2, 20)):  # Sample first 18 rows
            cell_val = ws.cell(row=r, column=c).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 40))
        ws.column_dimensions[get_column_letter(c)].width = max_len + 2


def main():
    print("Loading source workbook...")
    src_wb = openpyxl.load_workbook(SOURCE)

    print("Creating output workbook...")
    dst_wb = openpyxl.Workbook()
    # Remove default sheet
    dst_wb.remove(dst_wb.active)

    # Preserve original sheet order
    original_order = src_wb.sheetnames

    # Generate all data first
    print("Generating data...")
    count_per_month = {"jan": 80, "feb": 100, "mar": 120}
    total = sum(count_per_month.values())

    contactos = gen_contactos(total)
    leads_rows, leads_meta = gen_leads(count_per_month)
    interacciones = gen_interacciones(leads_meta)
    calificacion = gen_calificacion(leads_meta)
    deals = gen_deals(leads_meta)
    log_trans = gen_log(leads_meta, deals)

    # Headers for each generated table
    gen_data = {
        "dim_contactos": (
            ["id_contacto", "nombre", "apellido", "email", "telefono_1", "telefono_2",
             "empresa", "area", "pais", "ciudad", "empleados", "nivel", "fecha_creacion"],
            contactos
        ),
        "fact_leads": (
            ["id_lead", "id_contacto", "id_campana", "id_vendedor_sdr", "status",
             "calidad_contacto", "servicio_interes", "fecha_ingreso", "fecha_asignacion",
             "fecha_ultimo_contacto", "numero_toques", "tipo_seguimiento",
             "status_seguimiento", "notas"],
            leads_rows
        ),
        "fact_interacciones": (
            ["id_interaccion", "id_lead", "id_deal", "id_vendedor", "tipo_interaccion",
             "resultado", "numero_toque", "timestamp", "duracion_seg", "notas"],
            interacciones
        ),
        "fact_calificacion": (
            ["id_calificacion", "id_lead", "entendio_info_marketing", "mostro_interes_genuino",
             "necesidad_puntual", "perfil_adecuado", "necesita_decision_tercero",
             "tiene_presupuesto", "monto_presupuesto", "asociacion_industria",
             "region", "tipo_membresia", "fecha_calificacion"],
            calificacion
        ),
        "fact_deals": (
            ["id_deal", "id_lead", "id_contacto", "id_vendedor_ae", "id_producto",
             "status_venta", "proyeccion", "monto_proyeccion", "monto_apartado",
             "monto_cierre", "fecha_pase_ventas", "fecha_primer_contacto_ae",
             "fecha_cierre", "razon_perdida", "descuento_pct", "es_recompra",
             "es_cliente_activo", "producto_cierre", "fuente_origen",
             "tipo_transaccion", "notas_vendedor"],
            deals
        ),
        "log_transacciones": (
            ["id_log", "timestamp", "entidad", "id_entidad", "usuario",
             "campo_modificado", "valor_anterior", "valor_nuevo"],
            log_trans
        ),
    }

    # Build sheets in original order
    print("Building sheets...")
    for sheet_name in original_order:
        if sheet_name in COPY_SHEETS:
            print(f"  Copying: {sheet_name}")
            dst_ws = dst_wb.create_sheet(title=sheet_name)
            copy_sheet(src_wb[sheet_name], dst_ws)
        elif sheet_name in gen_data:
            headers, data = gen_data[sheet_name]
            print(f"  Generating: {sheet_name} ({len(data)} rows)")
            dst_ws = dst_wb.create_sheet(title=sheet_name)
            write_sheet(dst_ws, headers, data)

    src_wb.close()

    # Save
    print(f"\nSaving to: {OUTPUT}")
    dst_wb.save(OUTPUT)
    dst_wb.close()

    print(f"\nDone! File: {os.path.abspath(OUTPUT)}")
    print(f"\nData summary:")
    print(f"  dim_contactos:      {len(contactos)} rows")
    print(f"  fact_leads:         {len(leads_rows)} rows (Jan:80, Feb:100, Mar:120)")
    print(f"  fact_interacciones: {len(interacciones)} rows")
    print(f"  fact_calificacion:  {len(calificacion)} rows")
    print(f"  fact_deals:         {len(deals)} rows")
    print(f"  log_transacciones:  {len(log_trans)} rows")
    print(f"\nReference sheets copied as-is: {', '.join(COPY_SHEETS)}")


if __name__ == "__main__":
    main()
